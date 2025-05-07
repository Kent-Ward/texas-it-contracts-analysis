import pandas as pd
import glob
import os
import re
import openpyxl
from openpyxl.utils import get_column_letter

# === 📁 Define paths ===
solicitations_path = 'C:/Users/Khemb/Documents/TS_Project/data/raw/solicitations/'
no_solicitations_path = 'C:/Users/Khemb/Documents/TS_Project/data/raw/no_solicitations_dir/'
output_path = 'C:/Users/Khemb/Documents/TS_Project/data/processed/master_tracker.xlsx'

# === ✅ Relevant NIGP codes ===
relevant_codes = [
    '20800', '20810', '20811', '20820', '20821', '20837', '20880',
    '20900', '20954', '20957', '20960', '91551', '91579', '91596', '91597',
    '91598', '91829', '91830', '91871', '91890', '91895', '92000', '92003',
    '92005', '92006', '92007', '92014', '92021', '92035', '92040', '92045',
    '92046', '92047', '92048', '92064', '92065', '92085', '92091',
    '92416', '92435', '92460', '92464', '96130', '96269'
]

# === 🧼 Clean and standardize each DataFrame ===
def clean_dataframe(df, contract_type):
    # Normalize column names
    df.columns = [col.strip().lower().replace(' ', '_') for col in df.columns]

    # Add contract type column
    df['contract_type'] = contract_type

    # 🧠 Assign project_name based on folder type
    if contract_type.lower() == 'solicitation' and 'name' in df.columns:
        df['project_name'] = df['name']
    elif contract_type.lower() == 'no solicitation':
        if 'title' in df.columns:
            df['project_name'] = df['title']
        elif 'name' in df.columns:
            df['project_name'] = df['name']
        else:
            df['project_name'] = ''
    else:
        df['project_name'] = ''

        # ✅ Drop 'name' column if it exists to avoid duplication
    if 'name' in df.columns:
        df.drop(columns=['name'], inplace=True)


    # 🎯 Standardize 'status' column from any likely source
    for col in ['status', 'award_status', 'comments', 'award_type', 'justification']:
        if col in df.columns:
            df['status'] = df[col]
            break
    if 'status' not in df.columns:
        df['status'] = ''

        # 🛠 Ensure solicitation_id is treated as string to prevent Excel formatting issues
    if 'solicitation_id' in df.columns:
        df['solicitation_id'] = df['solicitation_id'].astype(str)

            

    # 🧹 Clean NIGP codes and filter by relevance
    if 'nigp_codes' in df.columns:
        df['nigp_codes'] = df['nigp_codes'].astype(str).str.replace(r'[\s-]', '', regex=True)
        df = df[df['nigp_codes'].apply(lambda x: any(code in x for code in relevant_codes))]
    else:
        print("⚠️ 'NIGP Codes' column not found. Skipping this sheet.")
        return pd.DataFrame()

    return df

# === 📊 Load & clean all Excel files from a folder ===
def load_and_combine_excels(path, contract_type_label):
    all_files = glob.glob(os.path.join(path, '*.xlsx'))
    print(f"🔍 Found {len(all_files)} Excel files in: {path}")

    df_list = []

    for file in all_files:
        try:
            xl = pd.ExcelFile(file)
            for sheet in xl.sheet_names:
                print(f"\n📂 Inspecting file: {file}\n  ➤ Sheet: {sheet}")
                df_temp = xl.parse(sheet)
                df_cleaned = clean_dataframe(df_temp, contract_type_label)
                if not df_cleaned.empty:
                    df_list.append(df_cleaned)
        except Exception as e:
            print(f"⚠️ Error processing {file}: {e}")

    print(f"✅ Loaded {len(df_list)} clean DataFrames for {contract_type_label}")
    return pd.concat(df_list, ignore_index=True) if df_list else pd.DataFrame()

# === 🚀 Run the full pipeline ===
solicitations_df = load_and_combine_excels(solicitations_path, 'Solicitation')
awards_df = load_and_combine_excels(no_solicitations_path, 'No Solicitation')

# Merge both datasets
master_tracker = pd.concat([solicitations_df, awards_df], ignore_index=True)


# Replace missing vendor names with a clear label
master_tracker['vendor_name'] = master_tracker['vendor_name'].fillna('Undisclosed')
master_tracker['vendor_name'] = master_tracker['vendor_name'].replace('', 'Undisclosed')

# Reorder key columns if they exist
cols_front = ['project_name', 'vendor_name', 'status', 'contract_type']
cols_reordered = cols_front + [col for col in master_tracker.columns if col not in cols_front]
master_tracker = master_tracker[cols_reordered]


# === Save initial Excel file (basic export)
master_tracker.to_excel(output_path, index=False, sheet_name='MasterTracker')

# ✅ Format solicitation_id to display correctly in Excel (no green triangle, no scientific notation)
if 'solicitation_id' in master_tracker.columns:
    master_tracker['solicitation_id'] = master_tracker['solicitation_id'].apply(
        lambda x: f"'{str(x)}" if pd.notnull(x) else x
    )

# === Open workbook to adjust solicitation_id formatting
wb = openpyxl.load_workbook(output_path)
ws = wb.active

# === Find column index for 'solicitation_id'
for col_idx, cell in enumerate(ws[1], start=1):
    if cell.value == 'solicitation_id':
        solicitation_col_index = col_idx
        break

# === Format as text + cast to string
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=solicitation_col_index)
    if cell.value is not None:
        cell.value = str(cell.value)
        cell.number_format = '@'

# === Save workbook (final)
wb.save(output_path)

# ✅ Final output message
print(f"\n✅ Master tracker created successfully at: {output_path}")
print(f"📊 Rows from Solicitation files: {len(solicitations_df)}")
print(f"📊 Rows from No Solicitation files: {len(awards_df)}")
print(f"📊 Total rows in Master Tracker: {len(master_tracker)}")
print("✅ solicitation_id column saved as Text — no more green triangle!")